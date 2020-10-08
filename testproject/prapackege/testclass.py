import openpyxl

book = openpyxl.load_workbook('C:\\Users\\BLW-LAPT-101\\Desktop\\pythonautomation\\mit.xlsx')

sheet = book.active
#cell = sheet.cell(row=, column=1)
#print(cell.value)

c1= sheet.cell(row=3,column=1)
c1.value="Amit"

book.save('C:\\Users\\BLW-LAPT-101\\Desktop\\pythonautomation\\mit.xlsx')